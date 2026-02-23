from __future__ import annotations

import logging
from datetime import datetime
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.services.models import Exhibit
from app.utils.errors import InfraError


logger = logging.getLogger(__name__)


_REQUIRED_COLUMNS = [
    "ID_Экспоната",
    "Название_экспоната",
    "Название_музея",
    "Описание",
    "Эпоха/Период",
    "Категория",
    "Дата_поступления",
    "Фото_файл",
    "История_связанная_с_Воронежем",
]


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value: Any) -> int:
    try:
        return int(str(value).strip())
    except Exception as exc:  # noqa: BLE001
        raise InfraError("Некорректный ID_Экспоната") from exc


def _to_date(value: Any, epoch) -> Optional[datetime.date]:
    if value is None or _to_str(value) == "":
        return None
    def _clean(s: str) -> str:
        return s.replace("г.", "").replace("г", "").replace("год", "").replace("Г.", "").strip()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, offset=epoch).date()
        except Exception:
            pass
    if isinstance(value, str) and value.strip().isdigit():
        try:
            return from_excel(float(value.strip()), offset=epoch).date()
        except Exception:
            pass
    if isinstance(value, str):
        value = _clean(value)
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:  # noqa: BLE001
            pass
    if isinstance(value, datetime):
        return value.date()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(_to_str(value), fmt).date()
        except ValueError:
            continue
    return None


def read_exhibits_from_xlsx(xlsx_path: str) -> list[Exhibit]:
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        epoch = wb.epoch

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [_to_str(h) for h in header_row]
        missing = [c for c in _REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise InfraError("В XLSX отсутствуют колонки: " + ", ".join(missing))

        idx = {name: headers.index(name) for name in _REQUIRED_COLUMNS}

        exhibits: list[Exhibit] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            raw_id = row[idx["ID_Экспоната"]]
            if raw_id is None or _to_str(raw_id) == "":
                continue

            exhibits.append(
                Exhibit(
                    exhibit_id=_to_int(raw_id),
                    name=_to_str(row[idx["Название_экспоната"]]),
                    museum=_to_str(row[idx["Название_музея"]]),
                    description=_to_str(row[idx["Описание"]]),
                    period=_to_str(row[idx["Эпоха/Период"]]),
                    category=_to_str(row[idx["Категория"]]),
                    received_date=_to_date(row[idx["Дата_поступления"]], epoch),
                    photo_file=_to_str(row[idx["Фото_файл"]]),
                    voronezh_story=_to_str(row[idx["История_связанная_с_Воронежем"]]),
                )
            )

        wb.close()
        return exhibits
    except InfraError:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.exception("Не удалось прочитать XLSX")
        raise InfraError("Ошибка чтения XLSX") from exc
